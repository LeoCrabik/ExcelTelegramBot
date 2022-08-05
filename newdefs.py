from openpyxl import Workbook,load_workbook
import math
from handlers import personal_actions as pa
from aiogram import types
from aiogram.types import ContentTypes
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
import os
from decimal import Decimal as d

date=None
TableI=0


def Nowdate():
	global date
	from datetime import datetime
	inDate = datetime.now()
	date_time = inDate.strftime("%m_%d_%Y_%H%M%S")
	date = date_time
	return(date_time)	

def Preresult(TON, TOB):
	try:
		global date
		try:
			wbSearch = Workbook()
			wbSearch = load_workbook(pa.Paths.FilePath)
			wsSearch = wbSearch.active
		except:
			print("Пизда 1")
		try:
			for i in range(1,10):
				for i1 in range(1,30):
					value=wsSearch.cell(row=i, column=i1).value
					if value == TOB:
						pa.ColumnsIDs.BalanceColumn = i1
						pa.ColumnsIDs.BalanceRaw = i
						break
		except: 
			print("Пизда 2")
		try:
			for i in range(1,10):
				for i1 in range(1,30):
					value=wsSearch.cell(row=i, column=i1).value
					if value == TON:
						pa.ColumnsIDs.NameColumn = i1
						pa.ColumnsIDs.NameRaw = i
						break
		except:
			print("Пизда 3")
		try:
			wbResult = Workbook()
			wsResult = wbResult.active
			i1=pa.ColumnsIDs.NameRaw-1
			i2=1
			value1 = wsSearch.cell(row= pa.ColumnsIDs.NameRaw, column = pa.ColumnsIDs.NameColumn).value
			value2 = wsSearch.cell(row= pa.ColumnsIDs.BalanceRaw, column = pa.ColumnsIDs.BalanceColumn).value
		except:
			print("Пизда 4")
		try:
			for i in range(pa.ColumnsIDs.BalanceRaw, wsSearch.max_row+1):
				value1 = wsSearch.cell(row=i, column=pa.ColumnsIDs.NameColumn).value 
				value2 = wsSearch.cell(row=i, column=pa.ColumnsIDs.BalanceColumn).value
				if wsSearch.cell(row=i, column=pa.ColumnsIDs.NameColumn).value is not None:
					wsResult.cell(row=i2, column=1).value=value1
					wsResult.cell(row=i2, column=2).value=value2
					i2+=1
		except:
			print("Пизда 5")
		try:
			wbEnd = Workbook()
			wsEnd = wbEnd.active
			wbResult.save(str(pa.Paths.PreresultFilePath))
		except:
			print("Пизда 6")
	except: print("Пизда")

def DataProcessing(Name):
	global TableI
	wbProcessing = Workbook()
	wbProcessing = load_workbook(pa.Paths.PreresultFilePath)
	wsProcessing = wbProcessing.active
	summary=0
	for i in range(1, wsProcessing.max_row+10): 
		if str(Name).strip()==str(wsProcessing.cell(row=i, column=1).value).strip():
			print(wsProcessing.cell(row=i, column=2).value)
			print()
			summary+=d(format(wsProcessing.cell(row=i, column=2).value, '.2f'))
			print(summary)
			print()
			print()
			#print(str(summary)+" "+str(wsProcessing.cell(row=i, column=1).value)+" "+str(wsProcessing.cell(row=i,column=2).value))
	pa.TableData.Name_list.append(str(Name))
	pa.TableData.Balance_list.append(summary)
	TableI+=1
	return(summary)
	summary=0

def DataProcessingFinal():
	global TableI
	wbFinal = Workbook()
	wsFinal = wbFinal.active
	for i in range (0, TableI):
		wsFinal.cell(row=i+1, column=1).value=str(pa.TableData.Name_list[i])
		wsFinal.cell(row=i+1, column=2).value=str(pa.TableData.Balance_list[i])	
	TableI=0
	wbEnd = Workbook()
	wsEnd = wbEnd.active
	wbFinal.save(str(pa.Paths.FinalFilePath))